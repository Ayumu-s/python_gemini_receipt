import subprocess
from PIL import Image
import os
import re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.styles import Border, Side
from datetime import datetime
import time
# geminiAPI用
import google.generativeai as genai

# レシート画像フォルダのパス
image_dir = "レシート画像"

# 出力フォーマットファイルを読み込み
with open("出力フォーマット.txt", "r", encoding="utf-8") as f:
    prompt = f.read()

def ask_gemini(prompt: str, image_path: str = None, mode: str = 'cli') -> str:
    if mode == 'cli':
        # 既存のCLI呼び出し
        cmd = [r"C:\Users\ayumu\AppData\Roaming\npm\gemini.cmd"]
        process = subprocess.Popen(
            cmd,
            stdin=subprocess.PIPE,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            encoding="utf-8"
        )
        if image_path:
            prompt = f"{prompt}\n@{image_path}"
        stdout, stderr = process.communicate(prompt)
        if stderr:
            print("エラー:", stderr)
        return stdout
    elif mode == 'api':
        # Gemini API呼び出し
        genai.configure(api_key="AIzaSyDAKcc74atvtwx1es-g-Kh6Gxi3rRPGn1s")
        model = genai.GenerativeModel('gemini-1.5-flash')
        if image_path:
            img = Image.open(image_path)
            response = model.generate_content([prompt, img])
        else:
            response = model.generate_content(prompt)
        return response.text
    else:
        raise ValueError("modeは 'cli' または 'api' を指定してください")

def extract_info_from_text(text):
    """テキストから日付、合計金額、画像名、勘定科目を抽出"""
    lines = text.split('\n')
    image_name = ""
    for line in lines:
        if "===" in line and ".jpg" in line:
            image_name = line.split("===")[1].strip().split("の分析結果")[0].strip()
            break
    date = ""
    for line in lines:
        if "日付：" in line:
            date = line.split("日付：")[1].strip()
            break
    account = ""
    for line in lines:
        if "勘定科目：" in line:
            account = line.split("勘定科目：")[1].strip()
            break
    total_amount = ""
    for line in lines:
        if "合計金額：" in line:
            total_amount = line.split("合計金額：")[1].strip()
            break
    return date, total_amount, image_name, account

def process_month_folder(month_folder):
    """月フォルダ内の画像を処理（エクセルD列重複チェック付き）"""
    month_path = os.path.join(image_dir, month_folder)
    if not os.path.exists(month_path):
        print(f"フォルダ {month_path} が存在しません")
        return []
    image_extensions = ['.jpg', '.jpeg', '.png', '.bmp', '.gif']
    image_files = [f for f in os.listdir(month_path) if os.path.splitext(f)[1].lower() in image_extensions]

    # ExcelのD列（画像名）を取得
    excel_file = "レシート、領収書.xlsx"
    existing_image_names = set()
    try:
        wb = load_workbook(excel_file)
        if month_folder in wb.sheetnames:
            ws = wb[month_folder]
            for row in ws.iter_rows(min_row=5, min_col=4, max_col=4):
                cell_value = row[0].value
                if cell_value:
                    existing_image_names.add(str(cell_value))
    except Exception as e:
        print(f"Excelファイルの読み込み中にエラーが発生しました: {e}")

    results = []
    for filename in image_files:
        if filename in existing_image_names:
            print(f"スキップ: {filename}（すでにExcelに存在）")
            continue
        image_path = os.path.join(month_path, filename)
        response_text = ask_gemini(prompt, image_path)
        # 先頭の空行を削除
        response_text = response_text.lstrip('\n')
        result_text = f"=== {filename} の分析結果 ===\n{response_text}"
        results.append(result_text)
        print(f"処理完了: {filename}")
    return results

def update_excel(month, results):
    """Excelファイルを更新"""
    excel_file = "レシート、領収書.xlsx"
    try:
        wb = load_workbook(excel_file)
        if month in wb.sheetnames:
            ws = wb[month]
        else:
            ws = wb.create_sheet(month)
        row = 5
        while ws.cell(row=row, column=2).value is not None:
            row += 1
        for result in results:
            date, total_amount, image_name, account = extract_info_from_text(result)
            ws.cell(row=row, column=2, value=date)
            if 'の分析結果 ===\n' in result:
                c_text = result.split('の分析結果 ===\n')[1]
            else:
                c_text = result
            ws.cell(row=row, column=3, value=c_text)
            # D列に画像名のハイパーリンクを設定
            if image_name:
                month_path = os.path.join(image_dir, month)
                image_path = os.path.abspath(os.path.join(month_path, image_name))
                cell = ws.cell(row=row, column=4, value=image_name)
                cell.hyperlink = image_path
                cell.style = "Hyperlink"
                thin = Side(border_style="thin", color="000000")
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            else:
                cell = ws.cell(row=row, column=4, value=image_name)
                thin = Side(border_style="thin", color="000000")
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            # E列に勘定科目を出力
            ws.cell(row=row, column=5, value=account)
            # F列に合計金額を出力
            total_num = re.sub(r'円', '', total_amount).replace(',', '').strip()
            try:
                total_num_value = int(total_num)
            except ValueError:
                total_num_value = None
            ws.cell(row=row, column=6, value=total_num_value)
            row += 1
        wb.save(excel_file)
        print(f"{month}のデータをExcelに保存しました")
    except Exception as e:
        print(f"Excelファイルの更新中にエラーが発生しました: {e}")

def main():
    month_folders = [f for f in os.listdir(image_dir) if os.path.isdir(os.path.join(image_dir, f))]
    for month_folder in month_folders:
        print(f"\n{month_folder}の処理を開始します...")
        results = process_month_folder(month_folder)
        if results:
            update_excel(month_folder, results)
            output_path = f"出力テキスト_{month_folder}.txt"
            with open(output_path, "w", encoding="utf-8") as f:
                for result in results:
                    f.write(result + "\n\n-----------------------------\n\n")
            print(f"{month_folder}の処理が完了しました")
        else:
            print(f"{month_folder}に処理対象の画像がありませんでした")

if __name__ == "__main__":
    start_time = time.time()
    main()
    end_time = time.time()
    elapsed = end_time - start_time
    print(f"処理時間: {elapsed:.2f}秒")