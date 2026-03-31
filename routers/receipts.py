import csv
import io
import mimetypes
import os
import re
import secrets
import time
import uuid
import zipfile
from datetime import date
from urllib.parse import quote

import openpyxl
from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, Request, UploadFile, status
from fastapi.responses import FileResponse, RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from openpyxl.styles import Alignment, Font, PatternFill
from PIL import Image, ImageOps, UnidentifiedImageError
import pillow_heif
pillow_heif.register_heif_opener()
from sqlalchemy import extract, func
from sqlalchemy.orm import Session

from database import get_db
from models import Receipt
from services.gemini import analyze_receipt

router = APIRouter()
templates = Jinja2Templates(directory="templates")

PRIVATE_UPLOAD_FOLDER = "storage/uploads"
LEGACY_UPLOAD_FOLDER = "static/uploads"
MAX_UPLOAD_BYTES = int(os.getenv("MAX_UPLOAD_BYTES", str(10 * 1024 * 1024)))
MAX_FILES_PER_UPLOAD = int(os.getenv("MAX_FILES_PER_UPLOAD", "10"))
LOGIN_RATE_LIMIT_WINDOW = 10 * 60
UPLOAD_RATE_LIMIT_WINDOW = 10 * 60
RATE_LIMIT_BUCKETS: dict[tuple[str, str], list[float]] = {}
# JPEGとして保存できるフォーマット（iOSのMPO・HEIFなど含む）
JPEG_COMPATIBLE_FORMATS = {"JPEG", "MPO", "HEIF", "HEIC", "TIFF", "BMP", "GIF"}
# PNG/WEBPはそのまま保存
PNG_FORMATS = {"PNG"}
WEBP_FORMATS = {"WEBP"}
APP_USERNAME = os.getenv("APP_USERNAME")
APP_PASSWORD = os.getenv("APP_PASSWORD")

if not APP_USERNAME or not APP_PASSWORD:
    raise RuntimeError("APP_USERNAME and APP_PASSWORD are required. Set them in .env before starting the app.")

os.makedirs(PRIVATE_UPLOAD_FOLDER, exist_ok=True)
os.makedirs(LEGACY_UPLOAD_FOLDER, exist_ok=True)


# ─────────────────────────── Security Helpers ───────────────────────────

def get_client_ip(request: Request) -> str:
    return request.client.host if request.client else "unknown"


def enforce_rate_limit(request: Request, scope: str, limit: int, window_seconds: int) -> None:
    now = time.time()
    key = (scope, get_client_ip(request))
    hits = RATE_LIMIT_BUCKETS.setdefault(key, [])
    hits[:] = [ts for ts in hits if now - ts < window_seconds]
    if len(hits) >= limit:
        raise HTTPException(status_code=status.HTTP_429_TOO_MANY_REQUESTS, detail="試行回数が多すぎます。しばらく待ってから再試行してください。")
    hits.append(now)


def get_csrf_token(request: Request) -> str:
    token = request.session.get("csrf_token")
    if not token:
        token = secrets.token_urlsafe(32)
        request.session["csrf_token"] = token
    return token


def validate_csrf(request: Request, csrf_token: str) -> None:
    session_token = request.session.get("csrf_token")
    if not csrf_token or not session_token or not secrets.compare_digest(csrf_token, session_token):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="不正なリクエストです。ページを再読み込みしてやり直してください。")


def render_template(request: Request, template_name: str, context: dict | None = None, status_code: int = 200):
    merged = dict(context or {})
    merged["csrf_token"] = get_csrf_token(request)
    return templates.TemplateResponse(request, template_name, merged, status_code=status_code)


def ensure_authenticated(request: Request) -> None:
    if not request.session.get("authenticated"):
        raise HTTPException(
            status_code=status.HTTP_303_SEE_OTHER,
            detail="ログインが必要です。",
            headers={"Location": "/login"},
        )


# ─────────────────────────── Upload Helpers ───────────────────────────

def safe_filename(filename: str) -> str:
    filename = os.path.basename(filename)
    filename = re.sub(r"[^\w\s.\-]", "", filename)
    return filename.strip() or "upload"


def build_content_disposition(filename: str, fallback_filename: str) -> str:
    return f'attachment; filename="{fallback_filename}"; filename*=UTF-8\'\'{quote(filename)}'


def normalize_image_bytes(contents: bytes) -> tuple[bytes, str]:
    if not contents:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="空のファイルはアップロードできません。")
    if len(contents) > MAX_UPLOAD_BYTES:
        raise HTTPException(status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE, detail=f"ファイルサイズは {MAX_UPLOAD_BYTES // (1024 * 1024)}MB 以下にしてください。")

    try:
        with Image.open(io.BytesIO(contents)) as image:
            source_format = (image.format or "").upper()

            # Pillowで開けた画像はすべて受け入れ、形式に応じてJPEG/PNG/WEBPに変換
            if source_format in PNG_FORMATS:
                save_format, extension = "PNG", ".png"
            elif source_format in WEBP_FORMATS:
                save_format, extension = "WEBP", ".webp"
            else:
                # JPEG互換（MPO, HEIF, BMP, TIFFなどiOS/Android各種）→ JPEG
                save_format, extension = "JPEG", ".jpg"

            normalized = ImageOps.exif_transpose(image)

            if save_format == "JPEG" and normalized.mode not in ("RGB", "L"):
                normalized = normalized.convert("RGB")
            elif save_format in {"PNG", "WEBP"} and normalized.mode == "P":
                normalized = normalized.convert("RGBA")

            # 長辺1600px以内にリサイズ（Gemini処理高速化 & DB節約）
            if max(normalized.width, normalized.height) > 1600:
                normalized.thumbnail((1600, 1600), Image.LANCZOS)

            output = io.BytesIO()
            if save_format == "JPEG":
                normalized.save(output, format="JPEG", quality=85, optimize=True)
            elif save_format == "PNG":
                normalized.save(output, format="PNG", optimize=True)
            elif save_format == "WEBP":
                normalized.save(output, format="WEBP", quality=85, method=6)

            return output.getvalue(), extension
    except UnidentifiedImageError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="画像ファイルとして認識できませんでした。別の画像をお試しください。") from exc


def make_stored_filename(extension: str) -> str:
    return f"{uuid.uuid4().hex}{extension}"


def resolve_receipt_file_path(receipt: Receipt | dict) -> str | None:
    stored_filename = receipt["stored_filename"] if isinstance(receipt, dict) else receipt.stored_filename
    original_filename = receipt["filename"] if isinstance(receipt, dict) else receipt.filename

    candidates: list[str] = []
    if stored_filename:
        candidates.append(os.path.join(PRIVATE_UPLOAD_FOLDER, stored_filename))
    if original_filename:
        candidates.append(os.path.join(PRIVATE_UPLOAD_FOLDER, original_filename))
        candidates.append(os.path.join(LEGACY_UPLOAD_FOLDER, original_filename))

    for candidate in candidates:
        if os.path.isfile(candidate):
            return candidate
    return None


def get_receipt_or_404(db: Session, receipt_id: int) -> Receipt:
    receipt = db.query(Receipt).filter(Receipt.id == receipt_id).first()
    if not receipt:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="レシートが見つかりません。")
    return receipt


# ─────────────────────────── Receipt Helpers ───────────────────────────

def get_receipt_reference_date(receipt_data: dict) -> date:
    return receipt_data["receipt_date"] or receipt_data["uploaded_at"].date()


def ensure_unique_path(path: str, used_paths: set[str]) -> str:
    base, ext = os.path.splitext(path)
    candidate = path
    counter = 2
    while candidate in used_paths:
        candidate = f"{base}_{counter}{ext}"
        counter += 1
    used_paths.add(candidate)
    return candidate


def extract_receipt_date(result: str, upload_year: int) -> date | None:
    match = re.search(r"日付：(\d{4})年(\d{1,2})月(\d{1,2})日", result)
    if match:
        try:
            return date(int(match.group(1)), int(match.group(2)), int(match.group(3)))
        except ValueError:
            pass
    match = re.search(r"日付：(\d{1,2})月(\d{1,2})日", result)
    if match:
        month, day = int(match.group(1)), int(match.group(2))
        year = upload_year - 1 if month > date.today().month else upload_year
        try:
            return date(year, month, day)
        except ValueError:
            pass
    return None


def parse_receipt_fields(result: str) -> dict:
    def get(pattern):
        match = re.search(pattern, result)
        return match.group(1).strip() if match else "－"

    return {
        "date_str": get(r"日付：(.+)"),
        "store": get(r"お店、会社名：(.+)"),
        "category": get(r"勘定科目：(.+)"),
        "total": get(r"合計金額：(.+)"),
    }


def parse_amount(total_str: str) -> int | None:
    match = re.search(r"[\d,]+", total_str)
    if match:
        try:
            return int(match.group().replace(",", ""))
        except ValueError:
            pass
    return None


def build_receipt_data(receipts_raw: list[Receipt]) -> list[dict]:
    result: list[dict] = []
    for receipt in receipts_raw:
        fields = parse_receipt_fields(receipt.result)
        reference_date = receipt.receipt_date or receipt.uploaded_at.date()
        result.append({
            "id": receipt.id,
            "filename": receipt.filename,
            "stored_filename": receipt.stored_filename,
            "uploaded_at": receipt.uploaded_at,
            "receipt_date": receipt.receipt_date,
            "result": receipt.result,
            "month_group": f"{reference_date.year}年{reference_date.month}月",
            "is_expense": receipt.is_expense if receipt.is_expense is not None else True,
            "total_int": parse_amount(fields["total"]),
            **fields,
        })
    return result


def compute_totals(receipts_data: list[dict]) -> tuple[str, dict]:
    year_sum = 0
    month_sums: dict[str, int] = {}
    for receipt in receipts_data:
        if not receipt["is_expense"]:
            continue
        amount = receipt["total_int"]
        if amount is not None:
            year_sum += amount
            month_sums[receipt["month_group"]] = month_sums.get(receipt["month_group"], 0) + amount

    def format_currency(value: int) -> str:
        return f"¥{value:,}"

    return format_currency(year_sum), {k: format_currency(v) for k, v in month_sums.items()}


def get_available_years(db: Session) -> list[int]:
    rows = (
        db.query(func.extract("year", Receipt.receipt_date).label("y"))
        .filter(Receipt.receipt_date.isnot(None))
        .distinct()
        .all()
    )
    return sorted([int(row.y) for row in rows if row.y], reverse=True)


# ─────────────────────────── Auth ───────────────────────────

@router.get("/login")
async def login_page(request: Request):
    if request.session.get("authenticated"):
        return RedirectResponse(url="/receipts", status_code=303)
    return render_template(request, "login.html")


@router.post("/login")
async def login(
    request: Request,
    username: str = Form(...),
    password: str = Form(...),
    csrf_token: str = Form(default=""),
):
    validate_csrf(request, csrf_token)
    enforce_rate_limit(request, "login", limit=5, window_seconds=LOGIN_RATE_LIMIT_WINDOW)

    if not (
        secrets.compare_digest(username, APP_USERNAME)
        and secrets.compare_digest(password, APP_PASSWORD)
    ):
        return render_template(
            request,
            "login.html",
            {"error": "ログイン情報が正しくありません。"},
            status_code=status.HTTP_401_UNAUTHORIZED,
        )

    request.session.clear()
    request.session["authenticated"] = True
    request.session["username"] = APP_USERNAME
    request.session["csrf_token"] = secrets.token_urlsafe(32)
    return RedirectResponse(url="/receipts", status_code=303)


@router.post("/logout")
async def logout(request: Request, csrf_token: str = Form(default="")):
    ensure_authenticated(request)
    validate_csrf(request, csrf_token)
    request.session.clear()
    return RedirectResponse(url="/login", status_code=303)


# ─────────────────────────── Pages ───────────────────────────

@router.get("/")
async def index(request: Request):
    ensure_authenticated(request)
    return render_template(request, "index.html")


@router.get("/upload")
async def upload_page(request: Request):
    ensure_authenticated(request)
    return render_template(
        request,
        "upload.html",
        {
            "max_files_per_upload": MAX_FILES_PER_UPLOAD,
            "max_upload_mb": MAX_UPLOAD_BYTES // (1024 * 1024),
        },
    )


@router.post("/upload")
async def upload(
    request: Request,
    files: list[UploadFile] = File(...),
    csrf_token: str = Form(default=""),
    db: Session = Depends(get_db),
):
    ensure_authenticated(request)
    validate_csrf(request, csrf_token)
    enforce_rate_limit(request, "upload", limit=20, window_seconds=UPLOAD_RATE_LIMIT_WINDOW)

    if not files:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="ファイルを選択してください。")
    if len(files) > MAX_FILES_PER_UPLOAD:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"同時アップロードは {MAX_FILES_PER_UPLOAD} 件までです。")

    errors = []
    for file in files:
        original_filename = safe_filename(file.filename or "upload")
        contents = await file.read()
        try:
            normalized_bytes, extension = normalize_image_bytes(contents)
        except HTTPException as exc:
            errors.append(f"{original_filename}: {exc.detail}")
            continue

        stored_filename = make_stored_filename(extension)
        mime = mimetypes.guess_type(f"x{extension}")[0] or "application/octet-stream"

        try:
            result = await analyze_receipt(normalized_bytes)
        except Exception:
            result = "AI解析に失敗しました。時間を置いて再度お試しください。"

        receipt_date = extract_receipt_date(result, date.today().year)
        db.add(
            Receipt(
                filename=original_filename,
                stored_filename=stored_filename,
                result=result,
                receipt_date=receipt_date,
                image_data=normalized_bytes,
                image_content_type=mime,
            )
        )

    if errors and not db.new:
        return render_template(
            request, "upload.html",
            {
                "error": "、".join(errors),
                "max_files_per_upload": MAX_FILES_PER_UPLOAD,
                "max_upload_mb": MAX_UPLOAD_BYTES // (1024 * 1024),
            },
            status_code=400,
        )

    db.commit()
    return RedirectResponse(url="/receipts", status_code=303)


@router.get("/receipts")
async def receipts_view(
    request: Request,
    db: Session = Depends(get_db),
    year: int = Query(default=None),
    month: int = Query(default=None),
    sort: str = Query(default="uploaded_at"),
):
    ensure_authenticated(request)
    query = db.query(Receipt)
    if year:
        query = query.filter(extract("year", Receipt.receipt_date) == year)
    if month:
        query = query.filter(extract("month", Receipt.receipt_date) == month)
    if sort == "receipt_date":
        query = query.order_by(Receipt.receipt_date.desc().nullslast())
    else:
        query = query.order_by(Receipt.uploaded_at.desc())

    receipts_data = build_receipt_data(query.all())
    year_total, month_totals = compute_totals(receipts_data)

    return render_template(
        request,
        "receipts.html",
        {
            "receipts": receipts_data,
            "available_years": get_available_years(db),
            "selected_year": year,
            "selected_month": month,
            "sort": sort,
            "year_total": year_total,
            "month_totals": month_totals,
        },
    )


@router.get("/receipts/{receipt_id}/image")
async def receipt_image(request: Request, receipt_id: int, db: Session = Depends(get_db)):
    ensure_authenticated(request)
    receipt = get_receipt_or_404(db, receipt_id)

    if receipt.image_data:
        from fastapi.responses import Response as _Response
        media_type = receipt.image_content_type or "application/octet-stream"
        return _Response(content=receipt.image_data, media_type=media_type)

    file_path = resolve_receipt_file_path(receipt)
    if not file_path:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="画像ファイルが見つかりません。")
    media_type = mimetypes.guess_type(file_path)[0] or "application/octet-stream"
    return FileResponse(file_path, media_type=media_type)


@router.post("/receipts/{receipt_id}/toggle-expense")
async def toggle_expense(
    receipt_id: int,
    request: Request,
    csrf_token: str = Form(default=""),
    db: Session = Depends(get_db),
):
    ensure_authenticated(request)
    validate_csrf(request, csrf_token)
    receipt = get_receipt_or_404(db, receipt_id)
    receipt.is_expense = not (receipt.is_expense if receipt.is_expense is not None else True)
    db.commit()
    return RedirectResponse(url="/receipts", status_code=303)


@router.get("/receipts/export")
async def export_receipts(
    request: Request,
    db: Session = Depends(get_db),
    year: int = Query(default=None),
    month: int = Query(default=None),
):
    ensure_authenticated(request)
    query = db.query(Receipt).order_by(Receipt.receipt_date.desc().nullslast())
    if year:
        query = query.filter(extract("year", Receipt.receipt_date) == year)
    if month:
        query = query.filter(extract("month", Receipt.receipt_date) == month)
    receipts_data = build_receipt_data(query.all())

    headers_row = [
        "レシート日付",
        "店舗名",
        "勘定科目",
        "合計金額",
        "経費対象",
        "ファイル名",
        "画像相対パス",
        "アップロード日時",
    ]

    receipt_objects = {r.id: r for r in db.query(Receipt).filter(Receipt.id.in_([r["id"] for r in receipts_data])).all()}

    used_image_paths: set[str] = set()
    export_rows = []
    for receipt in receipts_data:
        reference_date = get_receipt_reference_date(receipt)
        db_receipt = receipt_objects.get(receipt["id"])
        has_image = bool(
            (db_receipt and db_receipt.image_data)
            or resolve_receipt_file_path(receipt)
        )
        image_zip_path = None
        image_relative_path = "－"

        if has_image:
            image_zip_path = ensure_unique_path(
                f"images/{reference_date.year}/{reference_date.month:02d}/{safe_filename(receipt['filename'])}",
                used_image_paths,
            )
            image_relative_path = image_zip_path.replace("/", "\\")

        export_rows.append({
            **receipt,
            "image_zip_path": image_zip_path,
            "image_relative_path": image_relative_path,
            "_db_receipt": db_receipt,
        })

    def make_row(receipt: dict) -> list[str]:
        return [
            receipt["date_str"],
            receipt["store"],
            receipt["category"],
            receipt["total"],
            "○" if receipt["is_expense"] else "×",
            receipt["filename"],
            receipt["image_relative_path"],
            receipt["uploaded_at"].strftime("%Y/%m/%d %H:%M"),
        ]

    label = f"_{year}年" if year else "_全期間"
    fallback_label = f"_{year}" if year else "_all"
    if month:
        label += f"{month}月"
        fallback_label += f"_{month:02d}"

    package_root = f"receipts{fallback_label}"
    xlsx_name = f"receipts{fallback_label}.xlsx"
    csv_name = f"receipts{fallback_label}.csv"
    content_disposition = build_content_disposition(
        filename=f"receipts{label}.zip",
        fallback_filename=f"{package_root}.zip",
    )

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "レシート一覧"
    worksheet.append(headers_row)

    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    image_path_col_index = headers_row.index("画像相対パス") + 1
    for receipt in export_rows:
        worksheet.append(make_row(receipt))
        if receipt["image_zip_path"]:
            cell = worksheet.cell(row=worksheet.max_row, column=image_path_col_index)
            cell.hyperlink = receipt["image_relative_path"]
            cell.font = Font(color="0563C1", underline="single")

    for col in worksheet.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        worksheet.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    excel_output = io.BytesIO()
    workbook.save(excel_output)
    excel_output.seek(0)

    csv_output = io.StringIO()
    writer = csv.writer(csv_output)
    writer.writerow(headers_row)
    for receipt in export_rows:
        writer.writerow(make_row(receipt))

    zip_output = io.BytesIO()
    with zipfile.ZipFile(zip_output, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr(f"{package_root}/{xlsx_name}", excel_output.getvalue())
        archive.writestr(f"{package_root}/{csv_name}", csv_output.getvalue().encode("utf-8-sig"))
        for receipt in export_rows:
            if not receipt["image_zip_path"]:
                continue
            image_bytes: bytes | None = None
            db_receipt = receipt.get("_db_receipt")
            if db_receipt and db_receipt.image_data:
                image_bytes = db_receipt.image_data
            if image_bytes is None:
                local_path = resolve_receipt_file_path(receipt)
                if local_path:
                    with open(local_path, "rb") as img_f:
                        image_bytes = img_f.read()
            if image_bytes:
                archive.writestr(f"{package_root}/{receipt['image_zip_path']}", image_bytes)

    zip_output.seek(0)
    return StreamingResponse(
        zip_output,
        media_type="application/zip",
        headers={"Content-Disposition": content_disposition},
    )


@router.get("/summary")
async def summary_page(
    request: Request,
    db: Session = Depends(get_db),
    year: int = Query(default=None),
):
    ensure_authenticated(request)
    query = db.query(Receipt).filter(Receipt.is_expense.is_(True))
    if year:
        query = query.filter(extract("year", Receipt.receipt_date) == year)
    receipts_data = build_receipt_data(query.all())

    category_totals: dict[str, dict] = {}
    month_totals: dict[str, int] = {}
    for receipt in receipts_data:
        category = receipt["category"]
        amount = receipt["total_int"] or 0
        if category not in category_totals:
            category_totals[category] = {"count": 0, "total": 0}
        category_totals[category]["count"] += 1
        category_totals[category]["total"] += amount

        month_group = receipt["month_group"]
        month_totals[month_group] = month_totals.get(month_group, 0) + amount

    grand_total = sum(value["total"] for value in category_totals.values())

    category_list = sorted(
        [
            {
                "category": key,
                "count": value["count"],
                "total": value["total"],
                "total_fmt": f"¥{value['total']:,}",
                "pct": round(value["total"] / grand_total * 100, 1) if grand_total else 0,
            }
            for key, value in category_totals.items()
        ],
        key=lambda item: item["total"],
        reverse=True,
    )

    month_list = sorted(
        [{"month": key, "total": value, "total_fmt": f"¥{value:,}"} for key, value in month_totals.items()],
        key=lambda item: item["month"],
    )

    return render_template(
        request,
        "summary.html",
        {
            "category_list": category_list,
            "month_list": month_list,
            "grand_total": f"¥{grand_total:,}",
            "receipt_count": len(receipts_data),
            "available_years": get_available_years(db),
            "selected_year": year,
        },
    )


@router.post("/receipts/delete")
async def delete_receipts(
    request: Request,
    ids: list[int] = Form(default=[]),
    csrf_token: str = Form(default=""),
    db: Session = Depends(get_db),
):
    ensure_authenticated(request)
    validate_csrf(request, csrf_token)
    if ids:
        for receipt in db.query(Receipt).filter(Receipt.id.in_(ids)).all():
            file_path = resolve_receipt_file_path(receipt)
            if file_path and os.path.commonpath([os.path.abspath(file_path), os.path.abspath(PRIVATE_UPLOAD_FOLDER)]) == os.path.abspath(PRIVATE_UPLOAD_FOLDER):
                try:
                    os.remove(file_path)
                except FileNotFoundError:
                    pass
        db.query(Receipt).filter(Receipt.id.in_(ids)).delete(synchronize_session=False)
        db.commit()
    return RedirectResponse(url="/receipts", status_code=303)


@router.get("/receipts/{receipt_id}/edit")
async def edit_receipt_page(receipt_id: int, request: Request, db: Session = Depends(get_db)):
    ensure_authenticated(request)
    receipt = get_receipt_or_404(db, receipt_id)
    return render_template(request, "edit_receipt.html", {"receipt": receipt})


@router.post("/receipts/{receipt_id}/edit")
async def edit_receipt(
    receipt_id: int,
    request: Request,
    result: str = Form(...),
    csrf_token: str = Form(default=""),
    db: Session = Depends(get_db),
):
    ensure_authenticated(request)
    validate_csrf(request, csrf_token)
    receipt = get_receipt_or_404(db, receipt_id)
    receipt.result = result
    receipt.receipt_date = extract_receipt_date(result, receipt.uploaded_at.year)
    db.commit()
    return RedirectResponse(url="/receipts", status_code=303)
